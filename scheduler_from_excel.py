import csv
import calendar
from datetime import datetime, date, time
from collections import defaultdict
from openpyxl import load_workbook, Workbook

DEFAULT_MAX_HOURS_PER_WEEK = 40

INPUT_WORKBOOK = "pool_scheduler_input.xlsx"

OUTPUT_SCHEDULE_XLSX = "monthly_schedule.xlsx"
OUTPUT_SCHEDULE_CSV = "monthly_schedule.csv"
OUTPUT_HOURS_CSV = "employee_hours.csv"
OUTPUT_UNFILLED_CSV = "unfilled_shifts.csv"
OUTPUT_EMPLOYEE_SCHEDULE_CSV = "employee_schedule.csv"


def parse_date_string(s):
    return datetime.strptime(str(s).strip(), "%Y-%m-%d").date()


def parse_excel_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return parse_date_string(value)


def parse_time_string(value):
    if isinstance(value, datetime):
        return value.time().replace(second=0, microsecond=0)
    if isinstance(value, time):
        return value.replace(second=0, microsecond=0)

    value = str(value).strip()
    hour, minute = map(int, value.split(":"))
    return time(hour=hour, minute=minute)


def combine_date_time(day_obj, hhmm_or_time):
    if isinstance(hhmm_or_time, time):
        return datetime.combine(day_obj, hhmm_or_time)
    return datetime.combine(day_obj, parse_time_string(hhmm_or_time))


def date_str(day_obj):
    return day_obj.strftime("%Y-%m-%d")


def day_name(day_obj):
    return day_obj.strftime("%A")


def get_week_key(day_obj):
    iso = day_obj.isocalendar()
    return f"{iso.year}-W{iso.week:02d}"


def shifts_overlap(a_start, a_end, b_start, b_end):
    return a_start < b_end and b_start < a_end


def daterange_for_month(year, month):
    _, last_day = calendar.monthrange(year, month)
    for day_num in range(1, last_day + 1):
        yield date(year, month, day_num)


def role_can_fill(employee_role, role_needed):
    if role_needed == "headguard":
        return employee_role == "headguard"
    if role_needed == "lifeguard":
        return employee_role in ("lifeguard", "headguard")
    if role_needed == "slide_attendant":
        return employee_role == "slide_attendant"
    if role_needed == "manager":
        return employee_role == "manager"
    if role_needed == "cashier":
        return employee_role == "cashier"
    return False


def get_sheet_rows(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    output = []

    for row in rows[1:]:
        if row is None:
            continue
        row_dict = {}
        all_blank = True
        for i, header in enumerate(headers):
            value = row[i] if i < len(row) else None
            if value not in (None, ""):
                all_blank = False
            row_dict[header] = value
        if not all_blank:
            output.append(row_dict)

    return output


def load_input_workbook(filename):
    wb = load_workbook(filename, data_only=True)

    required_sheets = [
        "Employees",
        "Unavailability",
        "AllowedAssignments",
        "StaffingRequirements",
    ]

    for sheet_name in required_sheets:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Workbook is missing required sheet: {sheet_name}")

    employees = load_employees_sheet(wb["Employees"])
    load_unavailability_sheet(wb["Unavailability"], employees)
    load_allowed_assignments_sheet(wb["AllowedAssignments"], employees)
    staffing_requirements = load_staffing_requirements_sheet(wb["StaffingRequirements"])

    return employees, staffing_requirements


def load_employees_sheet(ws):
    employees = {}

    for row in get_sheet_rows(ws):
        employee_id = str(row.get("employee_id", "")).strip()
        if not employee_id:
            continue

        active_value = str(row.get("active", "1")).strip().lower()
        if active_value not in ("1", "true", "yes", "y"):
            continue

        name = str(row.get("name", "")).strip()
        role = str(row.get("role", "")).strip()

        max_hours_raw = row.get("max_hours_per_week", DEFAULT_MAX_HOURS_PER_WEEK)
        max_hours = float(max_hours_raw if max_hours_raw not in (None, "") else DEFAULT_MAX_HOURS_PER_WEEK)

        employees[employee_id] = {
            "employee_id": employee_id,
            "name": name,
            "role": role,
            "max_hours_per_week": max_hours,
            "unavailable_dates": set(),
            "allowed_assignments": "any",
        }

    return employees


def load_unavailability_sheet(ws, employees):
    for row in get_sheet_rows(ws):
        employee_id = str(row.get("employee_id", "")).strip()
        raw_date = row.get("date")

        if not employee_id or employee_id not in employees or raw_date in (None, ""):
            continue

        day = parse_excel_date(raw_date)
        employees[employee_id]["unavailable_dates"].add(date_str(day))


def load_allowed_assignments_sheet(ws, employees):
    temp_rules = defaultdict(list)

    for row in get_sheet_rows(ws):
        employee_id = str(row.get("employee_id", "")).strip()
        if not employee_id or employee_id not in employees:
            continue

        rule = {}

        pool = row.get("pool")
        shift_label = row.get("shift_label")
        days_of_week = row.get("days_of_week")

        if pool not in (None, ""):
            rule["pool"] = str(pool).strip()

        if shift_label not in (None, ""):
            rule["shift_label"] = str(shift_label).strip()

        if days_of_week not in (None, ""):
            rule["days_of_week"] = [
                d.strip() for d in str(days_of_week).split("|") if d.strip()
            ]

        if rule:
            temp_rules[employee_id].append(rule)

    for employee_id, rules in temp_rules.items():
        employees[employee_id]["allowed_assignments"] = rules


def load_staffing_requirements_sheet(ws):
    requirements = []

    for row in get_sheet_rows(ws):
        pool = row.get("pool")
        day_type = row.get("day_type")
        shift_label = row.get("shift_label")
        role_needed = row.get("role_needed")
        start_time = row.get("start_time")
        end_time = row.get("end_time")
        count_needed = row.get("count_needed")

        if any(v in (None, "") for v in [pool, day_type, shift_label, role_needed, start_time, end_time, count_needed]):
            continue

        requirements.append({
            "pool": str(pool).strip(),
            "day_type": str(day_type).strip(),
            "shift_label": str(shift_label).strip(),
            "role_needed": str(role_needed).strip(),
            "start_time": parse_time_string(start_time),
            "end_time": parse_time_string(end_time),
            "count_needed": int(count_needed),
        })

    return requirements


def requirement_applies_to_day(requirement, day_obj):
    weekday = day_obj.strftime("%a")
    day_type = requirement["day_type"]

    if day_type == "All":
        return True
    if day_type == weekday:
        return True
    if day_type == "Mon-Sat" and weekday in {"Mon", "Tue", "Wed", "Thu", "Fri", "Sat"}:
        return True
    return False


def assignment_matches_rule(shift, rule):
    if "pool" in rule and rule["pool"] != shift["pool"]:
        return False
    if "shift_label" in rule and rule["shift_label"] != shift["shift_label"]:
        return False
    if "days_of_week" in rule and day_name(shift["date"]) not in rule["days_of_week"]:
        return False
    return True


def employee_allowed_for_shift(emp, shift):
    if date_str(shift["date"]) in emp["unavailable_dates"]:
        return False

    allowed = emp["allowed_assignments"]
    if allowed == "any":
        return True

    return any(assignment_matches_rule(shift, rule) for rule in allowed)


def employee_restriction_score(emp):
    if emp["allowed_assignments"] == "any":
        score = 9999
    else:
        score = len(emp["allowed_assignments"])
    score += len(emp["unavailable_dates"]) * 0.01
    return score


def generate_monthly_shifts(year, month, staffing_requirements):
    shifts = []

    for day in daterange_for_month(year, month):
        for req in staffing_requirements:
            if not requirement_applies_to_day(req, day):
                continue

            start_dt = combine_date_time(day, req["start_time"])
            end_dt = combine_date_time(day, req["end_time"])
            duration_hours = (end_dt - start_dt).total_seconds() / 3600.0

            shifts.append({
                "date": day,
                "pool": req["pool"],
                "shift_label": req["shift_label"],
                "role_needed": req["role_needed"],
                "start": start_dt,
                "end": end_dt,
                "count_needed": req["count_needed"],
                "duration_hours": duration_hours,
            })

    return shifts


def expand_shift_slots(shifts):
    slots = []
    for shift in shifts:
        for i in range(shift["count_needed"]):
            slot = shift.copy()
            slot["slot_number"] = i + 1
            slots.append(slot)
    return slots


def build_schedule(year, month, employees_dict, staffing_requirements):
    employees = list(employees_dict.values())
    shifts = generate_monthly_shifts(year, month, staffing_requirements)
    slots = expand_shift_slots(shifts)

    assignments = []
    unfilled = []

    employee_assignments = defaultdict(list)
    employee_week_hours = defaultdict(lambda: defaultdict(float))
    employee_total_hours = defaultdict(float)

    role_priority = {
        "manager": 0,
        "cashier": 1,
        "slide_attendant": 2,
        "headguard": 3,
        "lifeguard": 4,
    }

    slots.sort(key=lambda s: (
        s["date"],
        s["start"],
        role_priority.get(s["role_needed"], 99),
        s["pool"],
    ))

    for slot in slots:
        eligible = []

        for emp in employees:
            if not role_can_fill(emp["role"], slot["role_needed"]):
                continue

            if not employee_allowed_for_shift(emp, slot):
                continue

            week_key = get_week_key(slot["date"])
            new_week_hours = employee_week_hours[emp["employee_id"]][week_key] + slot["duration_hours"]
            if new_week_hours > emp["max_hours_per_week"]:
                continue

            overlapping = False
            for a in employee_assignments[emp["employee_id"]]:
                if a["date"] == slot["date"] and shifts_overlap(slot["start"], slot["end"], a["start"], a["end"]):
                    overlapping = True
                    break
            if overlapping:
                continue

            eligible.append(emp)

        if not eligible:
            unfilled.append({
                "date": date_str(slot["date"]),
                "day": day_name(slot["date"]),
                "pool": slot["pool"],
                "shift_label": slot["shift_label"],
                "role_needed": slot["role_needed"],
                "start": slot["start"].strftime("%H:%M"),
                "end": slot["end"].strftime("%H:%M"),
                "duration_hours": slot["duration_hours"],
                "slot_number": slot["slot_number"],
            })
            continue

        def candidate_key(emp):
            week_key = get_week_key(slot["date"])
            headguard_penalty = 0
            if slot["role_needed"] == "lifeguard" and emp["role"] == "headguard":
                headguard_penalty = 1

            return (
                employee_restriction_score(emp),
                employee_week_hours[emp["employee_id"]][week_key],
                employee_total_hours[emp["employee_id"]],
                headguard_penalty,
                emp["name"],
                emp["employee_id"],
            )

        eligible.sort(key=candidate_key)
        chosen = eligible[0]

        assignment = {
            "employee_id": chosen["employee_id"],
            "employee": chosen["name"],
            "employee_role": chosen["role"],
            "date": slot["date"],
            "day": day_name(slot["date"]),
            "pool": slot["pool"],
            "shift_label": slot["shift_label"],
            "assigned_role": slot["role_needed"],
            "start": slot["start"],
            "end": slot["end"],
            "hours": slot["duration_hours"],
        }

        assignments.append(assignment)
        employee_assignments[chosen["employee_id"]].append(assignment)

        week_key = get_week_key(slot["date"])
        employee_week_hours[chosen["employee_id"]][week_key] += slot["duration_hours"]
        employee_total_hours[chosen["employee_id"]] += slot["duration_hours"]

    return assignments, unfilled, employee_week_hours, employee_total_hours


def export_schedule_csv(assignments, filename):
    rows = sorted(assignments, key=lambda a: (a["date"], a["start"], a["pool"], a["assigned_role"], a["employee"]))

    with open(filename, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow([
            "Date", "Day", "Pool", "Shift Label", "Assigned Role",
            "Employee ID", "Employee", "Employee Original Role", "Start", "End", "Hours"
        ])

        for a in rows:
            writer.writerow([
                date_str(a["date"]),
                a["day"],
                a["pool"],
                a["shift_label"],
                a["assigned_role"],
                a["employee_id"],
                a["employee"],
                a["employee_role"],
                a["start"].strftime("%H:%M"),
                a["end"].strftime("%H:%M"),
                a["hours"],
            ])


def export_unfilled_csv(unfilled, filename):
    rows = sorted(unfilled, key=lambda x: (x["date"], x["start"], x["pool"], x["role_needed"]))

    with open(filename, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow([
            "Date", "Day", "Pool", "Shift Label", "Role Needed",
            "Start", "End", "Hours", "Slot Number"
        ])

        for u in rows:
            writer.writerow([
                u["date"],
                u["day"],
                u["pool"],
                u["shift_label"],
                u["role_needed"],
                u["start"],
                u["end"],
                u["duration_hours"],
                u["slot_number"],
            ])


def export_hours_csv(employees_dict, employee_week_hours, employee_total_hours, filename):
    all_weeks = set()
    for _, week_map in employee_week_hours.items():
        for wk in week_map.keys():
            all_weeks.add(wk)
    all_weeks = sorted(all_weeks)

    employees = sorted(employees_dict.values(), key=lambda e: e["name"])

    with open(filename, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Employee ID", "Employee", "Role"] + all_weeks + ["Total Hours"])

        for emp in employees:
            row = [emp["employee_id"], emp["name"], emp["role"]]
            for wk in all_weeks:
                row.append(employee_week_hours[emp["employee_id"]].get(wk, 0))
            row.append(employee_total_hours.get(emp["employee_id"], 0))
            writer.writerow(row)


def export_employee_schedule_csv(assignments, filename):
    grouped = defaultdict(list)
    for a in assignments:
        grouped[a["employee_id"]].append(a)

    with open(filename, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow([
            "Employee ID", "Employee", "Date", "Day", "Pool", "Shift Label",
            "Assigned Role", "Start", "End", "Hours"
        ])

        for employee_id in sorted(grouped.keys()):
            person_assignments = sorted(grouped[employee_id], key=lambda a: (a["date"], a["start"]))
            for a in person_assignments:
                writer.writerow([
                    a["employee_id"],
                    a["employee"],
                    date_str(a["date"]),
                    a["day"],
                    a["pool"],
                    a["shift_label"],
                    a["assigned_role"],
                    a["start"].strftime("%H:%M"),
                    a["end"].strftime("%H:%M"),
                    a["hours"],
                ])


def auto_fit_worksheet_columns(ws):
    for col_cells in ws.columns:
        max_length = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_length:
                max_length = len(value)
        ws.column_dimensions[col_letter].width = min(max_length + 2, 30)


def export_schedule_workbook(assignments, unfilled, employees_dict, employee_week_hours, employee_total_hours, filename):
    wb = Workbook()

    ws_schedule = wb.active
    ws_schedule.title = "MonthlySchedule"

    ws_schedule.append([
        "Date", "Day", "Pool", "Shift Label", "Assigned Role",
        "Employee ID", "Employee", "Employee Original Role", "Start", "End", "Hours"
    ])

    rows = sorted(assignments, key=lambda a: (a["date"], a["start"], a["pool"], a["assigned_role"], a["employee"]))
    for a in rows:
        ws_schedule.append([
            date_str(a["date"]),
            a["day"],
            a["pool"],
            a["shift_label"],
            a["assigned_role"],
            a["employee_id"],
            a["employee"],
            a["employee_role"],
            a["start"].strftime("%H:%M"),
            a["end"].strftime("%H:%M"),
            a["hours"],
        ])

    ws_unfilled = wb.create_sheet("UnfilledShifts")
    ws_unfilled.append([
        "Date", "Day", "Pool", "Shift Label", "Role Needed",
        "Start", "End", "Hours", "Slot Number"
    ])

    unfilled_rows = sorted(unfilled, key=lambda x: (x["date"], x["start"], x["pool"], x["role_needed"]))
    for u in unfilled_rows:
        ws_unfilled.append([
            u["date"],
            u["day"],
            u["pool"],
            u["shift_label"],
            u["role_needed"],
            u["start"],
            u["end"],
            u["duration_hours"],
            u["slot_number"],
        ])

    ws_hours = wb.create_sheet("EmployeeHours")
    all_weeks = set()
    for _, week_map in employee_week_hours.items():
        for wk in week_map.keys():
            all_weeks.add(wk)
    all_weeks = sorted(all_weeks)

    ws_hours.append(["Employee ID", "Employee", "Role"] + all_weeks + ["Total Hours"])

    employees = sorted(employees_dict.values(), key=lambda e: e["name"])
    for emp in employees:
        row = [emp["employee_id"], emp["name"], emp["role"]]
        for wk in all_weeks:
            row.append(employee_week_hours[emp["employee_id"]].get(wk, 0))
        row.append(employee_total_hours.get(emp["employee_id"], 0))
        ws_hours.append(row)

    ws_employee_schedule = wb.create_sheet("EmployeeSchedule")
    ws_employee_schedule.append([
        "Employee ID", "Employee", "Date", "Day", "Pool", "Shift Label",
        "Assigned Role", "Start", "End", "Hours"
    ])

    grouped = defaultdict(list)
    for a in assignments:
        grouped[a["employee_id"]].append(a)

    for employee_id in sorted(grouped.keys()):
        person_assignments = sorted(grouped[employee_id], key=lambda a: (a["date"], a["start"]))
        for a in person_assignments:
            ws_employee_schedule.append([
                a["employee_id"],
                a["employee"],
                date_str(a["date"]),
                a["day"],
                a["pool"],
                a["shift_label"],
                a["assigned_role"],
                a["start"].strftime("%H:%M"),
                a["end"].strftime("%H:%M"),
                a["hours"],
            ])

    for ws in wb.worksheets:
        auto_fit_worksheet_columns(ws)

    wb.save(filename)


def print_summary(year, month, assignments, unfilled, employees_dict, employee_total_hours):
    print("=" * 60)
    print(f"SCHEDULE SUMMARY FOR {year}-{month:02d}")
    print("=" * 60)
    print(f"Employees entered: {len(employees_dict)}")
    print(f"Assignments made: {len(assignments)}")
    print(f"Unfilled slots:    {len(unfilled)}")
    print()

    role_counts = defaultdict(int)
    for emp in employees_dict.values():
        role_counts[emp["role"]] += 1

    print("Employees by role:")
    for role in sorted(role_counts.keys()):
        print(f"  {role}: {role_counts[role]}")
    print()

    print("Top 10 employees by scheduled hours:")
    top = sorted(employee_total_hours.items(), key=lambda x: (-x[1], x[0]))[:10]
    for employee_id, hours in top:
        name = employees_dict[employee_id]["name"]
        print(f"  {name} ({employee_id}): {hours:.2f}")
    print()

    if unfilled:
        print("First 15 unfilled slots:")
        for u in unfilled[:15]:
            print(
                f"  {u['date']} {u['pool']} {u['shift_label']} "
                f"{u['role_needed']} {u['start']}-{u['end']} slot #{u['slot_number']}"
            )
    else:
        print("No unfilled slots.")
    print()


def main():
    year = int(input("Enter year (example 2026): ").strip())
    month = int(input("Enter month number 1-12: ").strip())

    employees, staffing_requirements = load_input_workbook(INPUT_WORKBOOK)

    assignments, unfilled, employee_week_hours, employee_total_hours = build_schedule(
        year, month, employees, staffing_requirements
    )

    export_schedule_csv(assignments, OUTPUT_SCHEDULE_CSV)
    export_unfilled_csv(unfilled, OUTPUT_UNFILLED_CSV)
    export_hours_csv(employees, employee_week_hours, employee_total_hours, OUTPUT_HOURS_CSV)
    export_employee_schedule_csv(assignments, OUTPUT_EMPLOYEE_SCHEDULE_CSV)
    export_schedule_workbook(
        assignments,
        unfilled,
        employees,
        employee_week_hours,
        employee_total_hours,
        OUTPUT_SCHEDULE_XLSX,
    )

    print_summary(year, month, assignments, unfilled, employees, employee_total_hours)

    print(f"Created: {OUTPUT_SCHEDULE_XLSX}")
    print(f"Created: {OUTPUT_SCHEDULE_CSV}")
    print(f"Created: {OUTPUT_HOURS_CSV}")
    print(f"Created: {OUTPUT_UNFILLED_CSV}")
    print(f"Created: {OUTPUT_EMPLOYEE_SCHEDULE_CSV}")


if __name__ == "__main__":
    main()